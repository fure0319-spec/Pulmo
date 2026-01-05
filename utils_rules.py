
import os
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import load_workbook, Workbook
from filelock import FileLock

DEFAULT_XLSX = os.getenv("RULES_XLSX", "rules.xlsx")
LOCK_PATH = os.getenv("RULES_LOCK", DEFAULT_XLSX + ".lock")

REQUIRED_COLS = ["category", "name", "keywords", "advice"]

def _header_map(ws) -> Dict[str, int]:
    hm = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        hm[str(v).strip()] = col
    return hm

def load_rules(xlsx_path: str = DEFAULT_XLSX) -> List[Dict[str, Any]]:
    """Load rules from xlsx. Expected columns: category, name, keywords, advice."""
    if not os.path.exists(xlsx_path):
        # Create an empty starter workbook
        wb = Workbook()
        ws = wb.active
        for i, col in enumerate(REQUIRED_COLS, 1):
            ws.cell(row=1, column=i, value=col)
        wb.save(xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb.active
    hm = _header_map(ws)

    missing = [c for c in REQUIRED_COLS if c not in hm]
    if missing:
        raise ValueError(f"rules.xlsx에 필수 열이 없습니다: {missing}. (필요: {REQUIRED_COLS})")

    rules: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        category = ws.cell(row=r, column=hm["category"]).value
        name = ws.cell(row=r, column=hm["name"]).value
        if not category or not name:
            continue
        keywords_raw = ws.cell(row=r, column=hm["keywords"]).value
        advice_raw = ws.cell(row=r, column=hm["advice"]).value

        rules.append({
            "category": str(category).strip(),
            "name": str(name).strip(),
            "keywords": [k.strip() for k in str(keywords_raw or "").split(",") if k.strip()],
            "advice": str(advice_raw or ""),
        })
    return rules

def save_rules(rules: List[Dict[str, Any]], xlsx_path: str = DEFAULT_XLSX) -> None:
    """Overwrite xlsx with rules list. Uses a file lock to avoid concurrent writes."""
    with FileLock(LOCK_PATH, timeout=10):
        wb = Workbook()
        ws = wb.active
        for i, col in enumerate(REQUIRED_COLS, 1):
            ws.cell(row=1, column=i, value=col)

        for row_i, rule in enumerate(rules, start=2):
            ws.cell(row=row_i, column=1, value=rule.get("category", ""))
            ws.cell(row=row_i, column=2, value=rule.get("name", ""))
            ws.cell(row=row_i, column=3, value=",".join(rule.get("keywords", []) or []))
            ws.cell(row=row_i, column=4, value=rule.get("advice", ""))

        wb.save(xlsx_path)

def categories(rules: List[Dict[str, Any]]) -> List[str]:
    cats = sorted({r["category"] for r in rules if r.get("category")})
    return ["전체"] + cats

def filter_rules(
    rules: List[Dict[str, Any]],
    category: str = "전체",
    query: str = "",
) -> List[Dict[str, Any]]:
    out = rules
    if category and category != "전체":
        out = [r for r in out if r.get("category") == category]
    q = (query or "").strip().lower()
    if q:
        def blob(r):
            return f'{r.get("name","")} {",".join(r.get("keywords",[]))} {r.get("advice","")}'.lower()
        out = [r for r in out if q in blob(r)]
    # stable sort by name
    out = sorted(out, key=lambda r: (r.get("name","").lower(), r.get("category","").lower()))
    return out

def upsert_rule(rules: List[Dict[str, Any]], rule: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], bool]:
    """Return (new_rules, created?). Match by (category,name)."""
    cat = (rule.get("category") or "").strip()
    name = (rule.get("name") or "").strip()
    if not cat or not name:
        raise ValueError("category와 name은 필수입니다.")
    created = True
    new_rules = []
    replaced = False
    for r in rules:
        if r.get("category","").strip() == cat and r.get("name","").strip() == name and not replaced:
            new_rules.append(rule)
            replaced = True
            created = False
        else:
            new_rules.append(r)
    if not replaced:
        new_rules.append(rule)
    return new_rules, created

def delete_rule(rules: List[Dict[str, Any]], category: str, name: str) -> List[Dict[str, Any]]:
    cat = (category or "").strip()
    nm = (name or "").strip()
    return [r for r in rules if not (r.get("category","").strip()==cat and r.get("name","").strip()==nm)]
